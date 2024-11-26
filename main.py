 
import os 
import cv2 
import dlib 
import numpy as np 
import face_recognition 
import tensorflow as tf 
from tensorflow import keras 
from scipy.spatial import distance as dist 
import openpyxl 
import datetime 
 
 
 
# Function to format elapsed time as hours, minutes, and seconds 
def format_time(elapsed_seconds): 
    hours, remainder = divmod(elapsed_seconds, 3600) 
    minutes, seconds = divmod(remainder, 60) 
    return f"{int(hours)}h {int(minutes)}m {int(seconds)}s" 
 
# Function to load and preprocess images 
def load_and_preprocess_image(image_path, target_size=(100, 100)): 
    image = cv2.imread(image_path) 
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) 
    image = cv2.resize(image, target_size) 
    return image 
 
# Function to create or append to a person's database 
def create_or_append_database(person_name, dataset_dir): 
    person_dir = os.path.join(dataset_dir, person_name) 
    if not os.path.exists(person_dir): 
        os.makedirs(person_dir, exist_ok=True) 
    return person_dir 
 
# Function to recognize a face and return the person's name using Euclidean distance 
def recognize_face_euclidean(face_encoding, known_face_encodings, known_face_names, 
threshold=0.6): 
    min_distance = threshold 
    name = "Unknown" 
    for i, known_encoding in enumerate(known_face_encodings): 
        d = dist.euclidean(face_encoding, known_encoding) 
        if d < min_distance: 
            min_distance = d 
            name = known_face_names[i] 
    return name 
 
# Function to remove duplicate images in a dataset 
def remove_duplicate_images(dataset_dir): 
    for person_name in os.listdir(dataset_dir): 
        person_dir = os.path.join(dataset_dir, person_name) 
        if os.path.isdir(person_dir): 
            image_files = [f for f in os.listdir(person_dir) if f.endswith(".jpg") or f.endswith(".png")] 
            seen_images = set() 
            for image_file in image_files: 
                image_path = os.path.join(person_dir, image_file) 
                image_hash = hash(open(image_path, "rb").read()) 
                if image_hash in seen_images: 
                    # Remove the duplicate image 
                    os.remove(image_path) 
                    print(f"Removed duplicate image: {image_path}") 
                else: 
 
 
                    seen_images.add(image_hash) 
 
# Function to create or update an Excel sheet 
def create_or_update_excel(excel_file, name): 
    try: 
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") 
 
        if not os.path.isfile(excel_file): 
            workbook = openpyxl.Workbook() 
            worksheet = workbook.active 
            worksheet.title = "Attendance" 
            worksheet.append(["Name", "Total Duration"]) 
        else: 
            workbook = openpyxl.load_workbook(excel_file) 
            if "Attendance" not in workbook.sheetnames: 
                worksheet = workbook.create_sheet(title="Attendance") 
                worksheet.append(["Name", "Total Duration"]) 
            else: 
                worksheet = workbook["Attendance"] 
 
        worksheet.append([name, timestamp]) 
 
        workbook.save(excel_file) 
 
        print(f"Record saved in {excel_file} at {timestamp}") 
 
    except Exception as e: 
        print(f"An error occurred while saving the record: {str(e)}") 
 
# Initialize the face detector and shape predictor from dlib 
face_detector = dlib.get_frontal_face_detector() 
shape_predictor = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat") 
 
# Step 1: Create a directory to store the dataset 
dataset_dir = "dataset" 
os.makedirs(dataset_dir, exist_ok=True) 
 
# Step 2: Data Collection (Real-time) 
cap = cv2.VideoCapture(0) 
 
known_face_encodings = [] 
known_face_names = [] 
 
# Load the existing face encodings and names from the dataset 
for person_name in os.listdir(dataset_dir): 
    person_dir = os.path.join(dataset_dir, person_name) 
    if os.path.isdir(person_dir): 
 
 
        image_files = [f for f in os.listdir(person_dir) if f.endswith(".jpg") or f.endswith(".png")] 
        if image_files: 
            images = [face_recognition.load_image_file(os.path.join(person_dir, f)) for f in image_files] 
            # Collect face encodings only if faces are detected in the image 
            person_face_encodings = [] 
            for image in images: 
                face_locations = face_recognition.face_locations(image) 
                if face_locations: 
                    face_encoding = face_recognition.face_encodings(image, face_locations)[0] 
                    person_face_encodings.append(face_encoding) 
            known_face_encodings.extend(person_face_encodings) 
            known_face_names.extend([person_name] * len(person_face_encodings)) 
            print(f"Loaded {len(person_face_encodings)} face encodings for {person_name}") 
 
# Remove duplicate images from the database 
remove_duplicate_images(dataset_dir) 
 
# Function to create the CNN model 
def create_cnn_model(input_shape): 
    model = keras.Sequential([ 
        keras.layers.Conv2D(32, (3, 3), activation='relu', input_shape=input_shape), 
        keras.layers.MaxPooling2D((2, 2)), 
        keras.layers.Conv2D(64, (3, 3), activation='relu'), 
        keras.layers.MaxPooling2D((2, 2)), 
        keras.layers.Flatten(), 
        keras.layers.Dense(128, activation='relu'), 
        keras.layers.Dense(len(set(known_face_names)), activation='softmax')  # Output layer 
    ]) 
    model.compile(optimizer='adam', 
                  loss='sparse_categorical_crossentropy', 
                  metrics=['accuracy']) 
    return model 
 
# Step 3: Create the CNN model 
input_shape = (100, 100, 1)  # Adjust input shape as per your image size 
cnn_model = create_cnn_model(input_shape) 
 
# Load the trained weights (if available) 
weights_file = "cnn_model_weights.h5" 
if os.path.exists(weights_file): 
    cnn_model.load_weights(weights_file) 
 
# Additional variables for attendance tracking 
recognized_names = {}  # Dictionary to track recognized names 
candidate_start_time = {}  # Dictionary to track candidate's appearance start time 
 
# Initialize attendance data dictionary 
 
 
attendance_data = {} 
 
# Create or update an Excel sheet for attendance 
attendance_excel_file = "attendance.xlsx" 
 
if not os.path.isfile(attendance_excel_file): 
    workbook = openpyxl.Workbook() 
    worksheet = workbook.active 
    worksheet.title = "Attendance" 
    worksheet.append(["Name", "First Appearance", "Last Appearance", "Duration"]) 
else: 
    workbook = openpyxl.load_workbook(attendance_excel_file) 
    if "Attendance" not in workbook.sheetnames: 
        worksheet = workbook.create_sheet(title="Attendance") 
        worksheet.append(["Name", "First Appearance", "Last Appearance", "Duration"]) 
    else: 
        worksheet = workbook["Attendance"] 
 
# Create or update an Excel sheet for final attendance 
final_attendance_excel_file = "final_attendance.xlsx" 
 
# Main loop for real-time face recognition and attendance tracking 
while True: 
    ret, frame = cap.read() 
    gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY) 
 
    faces = face_detector(gray_frame) 
 
    for face in faces: 
        x, y, w, h = face.left(), face.top(), face.width(), face.height() 
        face_image = gray_frame[y:y + h, x:x + w] 
 
        if face_image.shape[0] > 0 and face_image.shape[1] > 0: 
            face_image = cv2.resize(face_image, (100, 100)) 
        else: 
            continue 
 
        # Make predictions using the CNN model 
        face_image = np.expand_dims(face_image, axis=0) 
        prediction = cnn_model.predict(face_image) 
 
        if prediction.shape[0] > 0:  # Check if predictions were made 
            predicted_label = np.argmax(prediction) 
            confidence = np.max(prediction) 
 
            # Get the face encoding of the detected face 
            face_locations = face_recognition.face_locations(frame) 
 
 
            face_encodings = face_recognition.face_encodings(frame, face_locations) 
 
            if face_encodings: 
                face_encoding = face_encodings[0]  # Assuming only one face is detected 
 
                # Recognize face using Euclidean distance 
                name = recognize_face_euclidean(face_encoding, known_face_encodings, 
known_face_names, threshold=0.6) 
 
                print(f"Recognized: {name} (Confidence: {confidence:.2f})") 
 
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2) 
                cv2.putText(frame, f"{name} ({confidence:.2f})", (x, y - 10), 
cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2) 
 
                # Update recognized names and attendance data 
                if name != "Unknown": 
                    if name in recognized_names: 
                        current_time = cv2.getTickCount() 
                        elapsed_time = (current_time - candidate_start_time[name]) / cv2.getTickFrequency() 
 
                        if elapsed_time >= 60.0:  # 1 minute 
                            if name not in attendance_data: 
                                attendance_data[name] = 0 
                            attendance_data[name] += elapsed_time 
                            recognized_names.pop(name) 
                            create_or_update_excel(attendance_excel_file, name)  # Save attendance record 
                            print(f"{name} appeared for {format_time(attendance_data[name])}") 
 
                    else: 
                        recognized_names[name] = True 
                        candidate_start_time[name] = cv2.getTickCount() 
 
    cv2.imshow("Face Recognition", frame) 
 
    if cv2.waitKey(1) & 0xFF == ord('q'): 
        break 
 
# Save final attendance data to the Excel sheet 
for name, total_time in attendance_data.items(): 
    worksheet.append([name, format_time(total_time)]) 
 
# Save the final attendance Excel file 
workbook.save(final_attendance_excel_file) 
print(f"Final Attendance saved in {final_attendance_excel_file}") 
 
# Release the camera and close OpenCV windows 
 
 
cap.release() 
cv2.destroyAllWindows() 